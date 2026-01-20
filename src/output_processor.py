from datetime import datetime
from pathlib import Path
import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from sklearn.metrics import roc_auc_score, average_precision_score
from sklearn.metrics import roc_curve, precision_recall_curve

global_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
global_script_dir = Path(__file__).parent.resolve()
global_root_folder = os.path.abspath(os.path.join(global_script_dir, os.pardir))
global_output_folder = os.path.join(global_root_folder, "output")
global_session_folder = os.path.join(global_output_folder, global_timestamp)

class OutputProcessor:
    def __init__(self):
        self.data = []
        self.excel_path = os.path.join(global_session_folder, f"uniguardian_{global_timestamp}.xlsx")
        self.save_path = global_session_folder
        self.plot_path = os.path.join(global_session_folder, "plot")

    def add_data(self, data):
        self.data.append(data)

    def write_output(self):
        if not os.path.exists(self.save_path):
            os.makedirs(self.save_path)
        if not os.path.exists(self.plot_path):
            os.makedirs(self.plot_path)
        for row in self.data:
            self.write_to_excel(row)
        self.finalize()

    def write_to_excel(self, data):
        new_row = pd.DataFrame([data])

        # If file does not exist → create it with mask_level sheet
        if not os.path.exists(self.excel_path):
            with pd.ExcelWriter(
                    self.excel_path,
                    engine="openpyxl",
                    mode="w"
            ) as writer:
                new_row.to_excel(writer, sheet_name="mask_level", index=False)

            print("output_file", self.excel_path)
            return

        # File exists → append to mask_level
        with pd.ExcelWriter(
                self.excel_path,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="overlay"
        ) as writer:
            try:
                df = pd.read_excel(self.excel_path, sheet_name="mask_level")
                df = pd.concat([df, new_row], ignore_index=True)
            except ValueError:
                # Sheet does not exist yet
                df = new_row

            df.to_excel(writer, sheet_name="mask_level", index=False)

        print("output_file", self.excel_path)

    def write_prompt_summary(self):
        df = pd.read_excel(self.excel_path, sheet_name="mask_level")

        grouped = df.groupby("prompt_id")

        rows = []
        for pid, g in grouped:
            idx_max = g["z_i"].idxmax()
            top_row = g.loc[idx_max]

            row = {
                "prompt_id": pid,
                "label": int(top_row["label"]),
                "num_masks": len(g),
                "z_max": float(top_row["z_max"]),
                "mean_S": float(top_row["mean_S"]),
                "std_S": float(top_row["std_S"]),
                "threshold": float(top_row["threshold"]),
                "is_poisoned": bool(top_row["is_poisoned"]),
                "is_correct": bool(
                    (top_row["label"] == 1 and top_row["is_poisoned"]) or
                    (top_row["label"] == 0 and not top_row["is_poisoned"])
                ),
                "top_mask_id": int(top_row["mask_id"]),
                "top_masked_words": top_row["masked_words"],
            }
            rows.append(row)

        summary_df = pd.DataFrame(rows)

        with pd.ExcelWriter(
                self.excel_path,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="replace",
        ) as writer:
            summary_df.to_excel(writer, sheet_name="prompt_summary", index=False)

    def write_overall_summary(self, excel_path):
        df = pd.read_excel(excel_path, sheet_name="prompt_summary")

        TP = ((df["label"] == 1) & (df["is_poisoned"] == True)).sum()
        FP = ((df["label"] == 0) & (df["is_poisoned"] == True)).sum()
        FN = ((df["label"] == 1) & (df["is_poisoned"] == False)).sum()
        TN = ((df["label"] == 0) & (df["is_poisoned"] == False)).sum()

        num_poisoned = (df["label"] == 1).sum()
        num_clean = (df["label"] == 0).sum()

        summary = {
            "num_prompts": len(df),
            "num_poisoned": num_poisoned,
            "num_clean": num_clean,
            "TP": TP,
            "FP": FP,
            "FN": FN,
            "TN": TN,
            "TPR (Recall)": TP / max(num_poisoned, 1),
            "FPR": FP / max(num_clean, 1),
            "Accuracy": (TP + TN) / max(len(df), 1),
        }

        summary_df = pd.DataFrame([summary])

        with pd.ExcelWriter(
                excel_path,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="replace",
        ) as writer:
            summary_df.to_excel(writer, sheet_name="overall_summary", index=False)

    def plot_zscore_distribution(self, df):
        plt.figure()
        for label, group in df.groupby("label"):
            plt.hist(
                group["z_max"],
                bins=30,
                alpha=0.6,
                label=f"label={label}"
            )
        z_dist_path = os.path.join(self.plot_path, "z_distribution.png")
        plt.axvline(df["threshold"].iloc[0], linestyle="--")
        plt.xlabel("Max Z-score")
        plt.ylabel("Count")
        plt.title("Z-score Distribution")
        plt.legend()
        plt.tight_layout()
        plt.savefig(z_dist_path)
        plt.close()
        return z_dist_path

    def plot_detection_scatter(self, df):
        plt.figure()
        scatter_path = os.path.join(self.plot_path, "z_scatter.png")
        colors = df["label"].map({0: "blue", 1: "red"})
        plt.scatter(df["prompt_id"], df["z_max"], c=colors)
        plt.axhline(df["threshold"].iloc[0], linestyle="--")
        plt.xlabel("Prompt ID")
        plt.ylabel("Max Z-score")
        plt.title("Max-Z Detection Result")
        plt.tight_layout()
        plt.savefig(scatter_path)
        plt.close()
        return scatter_path

    def plot_confusion(self, summary):
        labels = ["TP", "FP", "FN", "TN"]
        values = [summary[k] for k in labels]
        conf_path = os.path.join(self.plot_path, "confusion.png")
        plt.figure()
        plt.bar(labels, values)
        plt.title("Detection Outcome Summary")
        plt.tight_layout()
        plt.savefig(conf_path)
        plt.close()
        return conf_path

    def insert_plot_sheet(self, plot_files):
        wb = load_workbook(self.excel_path)

        if "plots" in wb.sheetnames:
            ws = wb["plots"]
        else:
            ws = wb.create_sheet("plots")

        row = 1
        for title, path in plot_files.items():
            ws.cell(row=row, column=1, value=title)
            img = Image(path)
            img.anchor = f"A{row + 1}"
            ws.add_image(img)
            row += 20

        wb.save(self.excel_path)

    def compute_auc_metrics(self, df):
        y_true = df["label"].values
        y_score = df["z_max"].values

        auroc = roc_auc_score(y_true, y_score)
        auprc = average_precision_score(y_true, y_score)

        return auroc, auprc

    def plot_roc(self, df):
        y_true = df["label"].values
        y_score = df["z_max"].values

        fpr, tpr, _ = roc_curve(y_true, y_score)
        auroc = roc_auc_score(y_true, y_score)
        roc_path = os.path.join(self.plot_path, "roc.png")
        plt.figure()
        plt.plot(fpr, tpr, label=f"AUROC = {auroc:.4f}")
        plt.plot([0, 1], [0, 1], linestyle="--")
        plt.xlabel("False Positive Rate")
        plt.ylabel("True Positive Rate")
        plt.title("ROC Curve (Max-Z)")
        plt.legend()
        plt.tight_layout()
        plt.savefig(roc_path)
        plt.close()
        return roc_path

    def plot_prc(self, df):
        y_true = df["label"].values
        y_score = df["z_max"].values

        precision, recall, _ = precision_recall_curve(y_true, y_score)
        auprc = average_precision_score(y_true, y_score)
        auprc_path = os.path.join(self.plot_path, "auprc.png")
        plt.figure()
        plt.plot(recall, precision, label=f"AUPRC = {auprc:.4f}")
        plt.xlabel("Recall")
        plt.ylabel("Precision")
        plt.title("Precision–Recall Curve (Max-Z)")
        plt.legend()
        plt.tight_layout()
        plt.savefig(auprc_path)
        plt.close()
        return auprc_path

    def finalize(self):
        # 1. Build summaries
        self.write_prompt_summary()
        self.write_overall_summary(self.excel_path)

        # 2. Load summary data
        df_prompt = pd.read_excel(self.excel_path, sheet_name="prompt_summary")

        # 3. Compute AUC metrics
        auroc, auprc = self.compute_auc_metrics(df_prompt)

        # 4. Generate plots
        paths = {
            "Z-score Distribution": self.plot_zscore_distribution(df_prompt),
            "Detection Scatter": self.plot_detection_scatter(df_prompt),
            "ROC Curve": self.plot_roc(df_prompt),
            "PR Curve": self.plot_prc(df_prompt),
        }

        # 5. Insert plots into Excel
        self.insert_plot_sheet(paths)

        print(f"Finalized report → {self.excel_path}")
        print(f"AUROC={auroc:.4f}, AUPRC={auprc:.4f}")
